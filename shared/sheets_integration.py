"""
shared/sheets_integration.py — Google Drive + Sheets integration.

Lets the app read from / write to the user's Google Sheets directly so
they don't have to download / re-upload Excel files. Uses an OAuth2
desktop flow on first use (one-time consent), then caches a token in
config/sheets_token.json that's shared across all sheet-aware features.

Public API
----------
status(resources_path) -> dict
    {configured, has_token, scopes}

reauthorize(resources_path, port=8600) -> dict
    Run the desktop OAuth flow. Opens system browser, waits for consent,
    writes config/sheets_token.json. Idempotent.

list_spreadsheets(resources_path, query='') -> list[dict]
    Newest-first list of every spreadsheet the user can see in Drive
    (filtered by the optional name fragment).

get_tabs(resources_path, spreadsheet_id) -> list[dict]
    Tabs (sheets) inside a spreadsheet, with row/col counts.

read_sheet(resources_path, spreadsheet_id, tab_name, range_a1='') -> list[list]
    2-D array of cell values.

write_range(resources_path, spreadsheet_id, range_a1, values) -> dict
    Overwrite a range with the given 2-D array. Range form: "TabName!A2:B10".

update_cell(resources_path, spreadsheet_id, tab_name, row, col, value) -> dict
    1-indexed row + column.

ensure_column(resources_path, spreadsheet_id, tab_name, header) -> int
    Find a column in row 1 by exact header (case-insensitive). If not
    present, append it to the right and return the new 1-based column
    index. Used when an op needs a "Status" column to write into.

batch_update_status(resources_path, spreadsheet_id, tab_name, status_col,
                    row_values) -> dict
    Write many { row_index → status_text } entries in a single API call.

Notes
-----
- Scope set kept minimal: Drive metadata.readonly (to list files) plus
  Spreadsheets full read/write.
- All errors are logged via shared.logger.print and surfaced as a dict
  {success: False, message: ...} so callers don't have to wrap.
"""

from __future__ import annotations

import json
import os
import threading
from pathlib import Path

# Scopes:
#   drive.metadata.readonly  — list spreadsheet files in the user's Drive
#   spreadsheets             — full read/write on the chosen spreadsheets
SCOPES = [
    'https://www.googleapis.com/auth/drive.metadata.readonly',
    'https://www.googleapis.com/auth/spreadsheets',
]

TOKEN_FILENAME = 'sheets_token.json'
CRED_FILENAME = 'gdrive_credentials.json'   # reuse the existing OAuth client

# ─────────────────────────────────────────────────────────────────────────────
# Auth / service builders
# ─────────────────────────────────────────────────────────────────────────────

def _token_path(resources_path) -> Path:
    return Path(resources_path) / 'config' / TOKEN_FILENAME


def _cred_path(resources_path) -> Path:
    return Path(resources_path) / 'config' / CRED_FILENAME


_load_error: dict = {'msg': ''}   # populated when _load_creds fails

def _load_creds(resources_path):
    """Load + refresh the cached OAuth credentials. Returns Credentials
    object or None on any failure (with a human-readable reason cached
    in _load_error)."""
    _load_error['msg'] = ''
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
    except Exception as e:
        _load_error['msg'] = f'google libs missing: {type(e).__name__}: {e}'
        return None
    p = _token_path(resources_path)
    if not p.exists():
        _load_error['msg'] = 'No sheets token — click Connect Google Sheets first.'
        return None
    try:
        creds = Credentials.from_authorized_user_file(str(p), SCOPES)
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            p.write_text(creds.to_json(), encoding='utf-8')
        return creds
    except Exception as e:
        _load_error['msg'] = f'token load failed: {type(e).__name__}: {e}'
        return None


_build_error: dict = {'msg': ''}

def _drive(resources_path):
    """Build a Google Drive v3 service."""
    _build_error['msg'] = ''
    creds = _load_creds(resources_path)
    if not creds:
        return None
    try:
        from googleapiclient.discovery import build
        return build('drive', 'v3', credentials=creds, cache_discovery=False)
    except Exception as e:
        _build_error['msg'] = f'Drive build: {type(e).__name__}: {e}'
        return None


def _sheets(resources_path):
    """Build a Google Sheets v4 service."""
    _build_error['msg'] = ''
    creds = _load_creds(resources_path)
    if not creds:
        return None
    try:
        from googleapiclient.discovery import build
        return build('sheets', 'v4', credentials=creds, cache_discovery=False)
    except Exception as e:
        _build_error['msg'] = f'Sheets build: {type(e).__name__}: {e}'
        return None


def _why_no_service() -> str:
    """Human-readable reason the Drive/Sheets service couldn't be built.
    Falls back to a generic 'not authorised' if both error slots empty."""
    return (_build_error['msg'] or _load_error['msg']
            or 'Sheets not authorized — click Connect Google Sheets.')


# ─────────────────────────────────────────────────────────────────────────────
# Auth status / consent flow
# ─────────────────────────────────────────────────────────────────────────────

def status(resources_path) -> dict:
    p = _token_path(resources_path)
    cred = _cred_path(resources_path)
    has_token = p.exists()
    scopes_ok = False
    if has_token:
        try:
            data = json.loads(p.read_text('utf-8'))
            scopes = set(data.get('scopes') or [])
            scopes_ok = all(s in scopes for s in SCOPES)
        except Exception:
            pass
    return {
        'configured': has_token and scopes_ok and cred.exists(),
        'has_token': has_token,
        'has_credentials': cred.exists(),
        'scopes_ok': scopes_ok,
    }


def reauthorize(resources_path, port: int = 8600) -> dict:
    """Run the desktop OAuth flow for the Sheets + Drive scopes."""
    cred = _cred_path(resources_path)
    if not cred.exists():
        return {'success': False,
                'message': 'config/gdrive_credentials.json missing'}
    try:
        from google_auth_oauthlib.flow import InstalledAppFlow
    except Exception as e:
        return {'success': False,
                'message': f'google-auth-oauthlib not available: {e}'}
    try:
        flow = InstalledAppFlow.from_client_secrets_file(str(cred), SCOPES)
        creds = flow.run_local_server(port=port, prompt='consent', open_browser=True)
        _token_path(resources_path).write_text(creds.to_json(), encoding='utf-8')
        return {'success': True,
                'message': 'Google Sheets re-authorized successfully'}
    except Exception as e:
        return {'success': False, 'message': f'Re-auth failed: {e}'}


# ─────────────────────────────────────────────────────────────────────────────
# Drive: list spreadsheets
# ─────────────────────────────────────────────────────────────────────────────

def list_spreadsheets(resources_path, query: str = '', limit: int = 50) -> dict:
    drive = _drive(resources_path)
    if not drive:
        return {'success': False,
                'message': _why_no_service(),
                'sheets': []}
    try:
        q_parts = [
            "mimeType = 'application/vnd.google-apps.spreadsheet'",
            'trashed = false',
        ]
        if query.strip():
            # name contains, escape single quotes
            esc = query.replace("'", "\\'")
            q_parts.append(f"name contains '{esc}'")
        q = ' and '.join(q_parts)
        res = drive.files().list(
            q=q, orderBy='modifiedTime desc', pageSize=limit,
            fields='files(id, name, modifiedTime, owners(displayName, emailAddress))',
        ).execute()
        files = res.get('files') or []
        out = []
        for f in files:
            owners = f.get('owners') or []
            owner = owners[0].get('displayName', '') if owners else ''
            out.append({
                'id': f['id'],
                'name': f['name'],
                'modified': f.get('modifiedTime'),
                'owner': owner,
            })
        return {'success': True, 'sheets': out}
    except Exception as e:
        return {'success': False, 'message': f'Drive list failed: {e}', 'sheets': []}


# ─────────────────────────────────────────────────────────────────────────────
# Sheets: tabs / read / write / append / column helpers
# ─────────────────────────────────────────────────────────────────────────────

def get_tabs(resources_path, spreadsheet_id: str) -> dict:
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service(), 'tabs': []}
    try:
        meta = s.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields='properties.title, sheets.properties',
        ).execute()
        tabs = []
        for sh in meta.get('sheets') or []:
            p = sh.get('properties') or {}
            grid = p.get('gridProperties') or {}
            tabs.append({
                'id': p.get('sheetId'),
                'title': p.get('title'),
                'rows': grid.get('rowCount'),
                'cols': grid.get('columnCount'),
            })
        return {'success': True,
                'spreadsheet_title': meta.get('properties', {}).get('title'),
                'tabs': tabs}
    except Exception as e:
        return {'success': False,
                'message': f'Sheets read failed: {type(e).__name__}: {(str(e) or repr(e))[:200]}',
                'tabs': []}


def _quote_tab(tab: str) -> str:
    """Single-quote a tab name for A1 ranges (escape inner quotes)."""
    safe = tab.replace("'", "''")
    return f"'{safe}'"


def read_sheet(resources_path, spreadsheet_id: str, tab_name: str,
               range_a1: str = '') -> dict:
    """Read a 2-D array of values. range_a1 should be the cell range
    WITHOUT the tab prefix (e.g. 'A1:Z' or '' for the whole tab)."""
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    try:
        if range_a1:
            r = f'{_quote_tab(tab_name)}!{range_a1}'
        else:
            r = _quote_tab(tab_name)
        res = s.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id, range=r,
            valueRenderOption='UNFORMATTED_VALUE',
        ).execute()
        return {'success': True, 'values': res.get('values') or []}
    except Exception as e:
        return {'success': False, 'message': f'Read failed: {e}'}


def write_range(resources_path, spreadsheet_id: str, range_a1: str,
                values: list[list]) -> dict:
    """Overwrite *range_a1* (must include the tab name, e.g.
    "Sheet1!A2:B10") with the given 2-D array."""
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    try:
        body = {'values': values}
        res = s.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id, range=range_a1,
            valueInputOption='USER_ENTERED', body=body,
        ).execute()
        return {'success': True,
                'updated_cells': res.get('updatedCells', 0),
                'updated_range': res.get('updatedRange')}
    except Exception as e:
        return {'success': False, 'message': f'Write failed: {e}'}


def update_cell(resources_path, spreadsheet_id: str, tab_name: str,
                row: int, col: int, value) -> dict:
    """Set a single cell (row + col are 1-based)."""
    from openpyxl.utils import get_column_letter
    a1 = f'{_quote_tab(tab_name)}!{get_column_letter(col)}{row}'
    return write_range(resources_path, spreadsheet_id, a1, [[value]])


def append_rows(resources_path, spreadsheet_id: str, tab_name: str,
                values: list[list]) -> dict:
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    try:
        res = s.spreadsheets().values().append(
            spreadsheetId=spreadsheet_id,
            range=_quote_tab(tab_name),
            valueInputOption='USER_ENTERED',
            insertDataOption='INSERT_ROWS',
            body={'values': values},
        ).execute()
        return {'success': True,
                'updates': res.get('updates', {})}
    except Exception as e:
        return {'success': False, 'message': f'Append failed: {e}'}


def ensure_column(resources_path, spreadsheet_id: str, tab_name: str,
                  header: str) -> dict:
    """Find a column in row 1 by exact header (case-insensitive). If not
    present, append it to the next free column and return its 1-based
    index. Used by ops that need a 'Status' column to write into."""
    from openpyxl.utils import get_column_letter
    headers_res = read_sheet(resources_path, spreadsheet_id, tab_name, '1:1')
    if not headers_res.get('success'):
        return {'success': False, 'message': headers_res.get('message')}
    rows = headers_res.get('values') or []
    headers = rows[0] if rows else []
    header_l = header.strip().lower()
    for i, h in enumerate(headers, 1):
        if str(h or '').strip().lower() == header_l:
            return {'success': True, 'col': i, 'created': False}
    new_col = len(headers) + 1
    a1 = f'{_quote_tab(tab_name)}!{get_column_letter(new_col)}1'
    res = write_range(resources_path, spreadsheet_id, a1, [[header]])
    if not res.get('success'):
        return {'success': False, 'message': res.get('message')}
    return {'success': True, 'col': new_col, 'created': True}


def batch_update_status(resources_path, spreadsheet_id: str, tab_name: str,
                        status_col: int, row_values: dict[int, str]) -> dict:
    """Push many status-cell updates in a single API call.

    row_values: {row_index_1based: 'Live'|'Missing'|'Error'|...}
    """
    if not row_values:
        return {'success': True, 'updated': 0}
    s = _sheets(resources_path)
    if not s:
        return {'success': False, 'message': _why_no_service()}
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(status_col)
    data = []
    for row, val in row_values.items():
        a1 = f'{_quote_tab(tab_name)}!{col_letter}{int(row)}'
        data.append({'range': a1, 'values': [[val]]})
    try:
        res = s.spreadsheets().values().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={'valueInputOption': 'USER_ENTERED', 'data': data},
        ).execute()
        return {'success': True, 'updated': res.get('totalUpdatedCells', 0)}
    except Exception as e:
        return {'success': False, 'message': f'Batch update failed: {e}'}


# ─────────────────────────────────────────────────────────────────────────────
# Convenience helper for ops: read a column by exact header
# ─────────────────────────────────────────────────────────────────────────────

def read_column_by_header(resources_path, spreadsheet_id: str, tab_name: str,
                          header: str) -> dict:
    """Find a column by exact header (case-insensitive) in row 1, then
    return the values from row 2 onward as [(row_index_1based, value)].
    Empty cells are skipped."""
    res = read_sheet(resources_path, spreadsheet_id, tab_name)
    if not res.get('success'):
        return {'success': False, 'message': res.get('message')}
    rows = res.get('values') or []
    if not rows:
        return {'success': True, 'col': None, 'rows': []}
    headers = rows[0]
    header_l = header.strip().lower()
    col_idx = None
    for i, h in enumerate(headers):
        if str(h or '').strip().lower() == header_l:
            col_idx = i
            break
    if col_idx is None:
        return {'success': False, 'message': f"Header '{header}' not found",
                'available_headers': [str(h) for h in headers]}
    out = []
    for ri, row in enumerate(rows[1:], start=2):   # row 2 onward
        v = row[col_idx] if col_idx < len(row) else None
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        out.append((ri, s))
    return {'success': True, 'col': col_idx + 1, 'rows': out}
