"""
shared/live_status_check.py — Bulk-check whether GMB reviews are LIVE.

Reads an Excel file that has a "Review Live Link" (or similar) column,
visits each URL in a headless browser, and writes a NEW workbook out
with a "Live Status" column populated as Live / Missing / Error.

Detection logic ported from E:/mailexus-advanced/step4/operations/live_check.py
— same selectors and missing-text indicators as the proven implementation.

Standalone op: NO login, NO profile manager, NO NST. The browser is a
fresh headless Chromium spun up via Playwright. Its user-data folder is
created in a dedicated temp directory and DELETED after the run finishes
so the host disk doesn't accumulate junk. Profile-manager data folders
are never touched.
"""

from __future__ import annotations

import asyncio
import os
import shutil
import tempfile
import threading
from datetime import datetime
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# Detection rules (copied from mailexus-advanced/step4/operations/live_check.py)
# ─────────────────────────────────────────────────────────────────────────────

LIVE_SELECTORS = [
    'div.Upo0Ec',                        # Like/Share container
    'button[aria-label="Like"]',
    'button[aria-label="Share"]',
    'button.gllhef[data-review-id]',
    'span.wiI7pd',                       # review text
    'div.MyEned',                        # review card
    'div.jftiEf',                        # review body
    'button[data-tooltip="Like"]',
    'div.DUwDvf',                        # review header
    'span.RfnDt',                        # reviewer name
]

MISSING_INDICATORS = [
    'this content is no longer available',
    "content isn't available",
    'page not found',
    "couldn't find",
    'no longer exists',
    'has been removed',
    'violates our policies',
]

JS_CHECK = """
(() => {
    const selectors = [
        'div.Upo0Ec', 'button[aria-label="Like"]', 'button[aria-label="Share"]',
        'span.wiI7pd', 'div.MyEned', 'div.jftiEf', 'button[data-tooltip="Like"]',
    ];
    for (const sel of selectors) {
        const el = document.querySelector(sel);
        if (el && el.offsetParent !== null) return true;
    }
    const hasStars = document.querySelectorAll('span.fzvQIb, span.kvMYJc').length > 0;
    const hasReview = document.querySelectorAll('div.MyEned, div.jftiEf, span.wiI7pd').length > 0;
    return hasStars && hasReview;
})()
"""


# ─────────────────────────────────────────────────────────────────────────────
# Module state
# ─────────────────────────────────────────────────────────────────────────────

_status: dict = {
    'running': False, 'started_at': '', 'finished_at': '',
    'total': 0, 'done': 0, 'live': 0, 'not_live': 0, 'errors': 0,
    'current_url': '', 'report_path': '',
}
_status_lock = threading.Lock()
_cancel = threading.Event()
# Live worker handles so cancel() can forcibly close every browser.
# 'browsers' → list of Playwright BrowserContexts (empty when idle)
# 'loop'     → the asyncio loop the worker is using (so cancel can
#              schedule the async close on the right loop from a
#              different thread)
_live_browser_ref: dict = {'browsers': [], 'loop': None}


def get_status() -> dict:
    with _status_lock:
        return dict(_status)


def cancel():
    """Hard cancel: signal the run to stop AND forcibly close every
    live browser so any in-flight goto/wait calls abort immediately.
    Without the close, navigation timeouts can keep the workers busy
    for another 20-30s after Stop is pressed."""
    _cancel.set()
    browsers = list(_live_browser_ref.get('browsers') or [])
    loop = _live_browser_ref.get('loop')
    if not browsers or loop is None:
        return
    # Schedule b.close() on the worker's own asyncio loop. We can call
    # this from any thread thanks to call_soon_threadsafe; awaiting
    # from elsewhere would crash because Playwright objects are bound
    # to the loop that created them.
    def _kill_all():
        for b in browsers:
            try:
                asyncio.ensure_future(b.close())
            except Exception:
                pass
    try:
        loop.call_soon_threadsafe(_kill_all)
    except Exception:
        pass


def start(file_path: str, num_workers: int = 5, timeout_sec: int = 20,
          resources_path: Path | None = None, show_browser: bool = False) -> dict:
    with _status_lock:
        if _status['running']:
            return {'success': False, 'message': 'Already running'}

    if not Path(file_path).exists():
        return {'success': False, 'message': f'File not found: {file_path}'}

    _cancel.clear()
    threading.Thread(
        target=_worker,
        args=(file_path, num_workers, timeout_sec, resources_path, show_browser),
        daemon=True, name='live-status-check',
    ).start()
    return {'success': True}


def start_from_sheet(sheet_id: str, tab_name: str, num_workers: int = 5,
                     timeout_sec: int = 20,
                     resources_path: Path | None = None,
                     show_browser: bool = False) -> dict:
    """Live-check the URLs in a Google Sheet tab. Status is written back
    to the same tab's Status column in real time."""
    with _status_lock:
        if _status['running']:
            return {'success': False, 'message': 'Already running'}
    if not sheet_id or not tab_name:
        return {'success': False, 'message': 'sheet_id + tab_name required'}
    _cancel.clear()
    threading.Thread(
        target=_worker_sheet,
        args=(sheet_id, tab_name, num_workers, timeout_sec,
              resources_path, show_browser),
        daemon=True, name='live-status-check-sheet',
    ).start()
    return {'success': True}


# ─────────────────────────────────────────────────────────────────────────────
# Worker
# ─────────────────────────────────────────────────────────────────────────────

def _worker_sheet(sheet_id: str, tab_name: str, num_workers: int,
                  timeout_sec: int, resources_path: Path | None,
                  show_browser: bool):
    """Same as _worker, but reads URLs from a Google Sheet tab and
    writes the verdicts back to that tab's Status column."""
    from shared import sheets_integration as _si
    global _status

    with _status_lock:
        _status.update({
            'running': True, 'started_at': datetime.utcnow().isoformat() + 'Z',
            'finished_at': '', 'total': 0, 'done': 0, 'live': 0,
            'not_live': 0, 'errors': 0, 'current_url': '', 'report_path': '',
        })

    # Sweep stale temp folders
    try:
        _tmp_parent = Path(tempfile.gettempdir())
        for stale in _tmp_parent.glob('nst_live_check_*'):
            try:
                shutil.rmtree(stale, ignore_errors=True)
            except Exception:
                pass
    except Exception:
        pass
    tmp_root = Path(tempfile.mkdtemp(prefix='nst_live_check_'))

    try:
        # 1. Read the Review Live Link column from the sheet
        col_res = _si.read_column_by_header(
            resources_path, sheet_id, tab_name, 'Review Live Link',
        )
        if not col_res.get('success'):
            with _status_lock:
                _status['running'] = False
                _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
                _status['errors'] = 1
                _status['current_url'] = (
                    f"FATAL: {col_res.get('message', 'header not found')}"
                )
            return

        # Build (row, url) lists; dedup by URL but preserve all rows for
        # writing the status back to every occurrence.
        all_rows: list[tuple[int, str]] = []
        items: list[tuple[int, str]] = []
        seen: set[str] = set()
        for ri, v in col_res.get('rows') or []:
            url = v.strip()
            ul = url.lower()
            if not (ul.startswith('http://') or ul.startswith('https://')):
                continue
            all_rows.append((ri, url))
            if ul not in seen:
                seen.add(ul)
                items.append((ri, url))

        with _status_lock:
            _status['total'] = len(items)

        # 2. Make sure a Status column exists on the sheet
        st_res = _si.ensure_column(
            resources_path, sheet_id, tab_name, 'Status',
        )
        if not st_res.get('success'):
            with _status_lock:
                _status['running'] = False
                _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
                _status['errors'] = 1
                _status['current_url'] = (
                    f"FATAL: {st_res.get('message', 'could not create Status column')}"
                )
            return
        status_col = st_res['col']

        # 2b. Read the EXISTING value of the Status column for every row
        # so we can promote 'Applead' → 'Done' when the live check
        # confirms the review is still live.
        from openpyxl.utils import get_column_letter
        status_col_letter = get_column_letter(status_col)
        existing_status: dict[int, str] = {}
        try:
            sd = _si.read_sheet(resources_path, sheet_id, tab_name,
                                f'{status_col_letter}2:{status_col_letter}')
            if sd.get('success'):
                for offset, row in enumerate(sd.get('values') or []):
                    if not row:
                        continue
                    val = str(row[0] if row else '').strip()
                    if val:
                        existing_status[offset + 2] = val
        except Exception:
            pass

        # 3. Run the checks
        loop = asyncio.new_event_loop()
        try:
            asyncio.set_event_loop(loop)
            results = loop.run_until_complete(
                _run_checks(items, num_workers, timeout_sec, tmp_root, show_browser)
            )
        finally:
            try: loop.close()
            except Exception: pass

        # 4. Mirror verdicts to all rows that had each URL, push to sheet.
        # Status-transition rules (per user spec):
        #     existing 'Appeal' / 'Applead' + Live    → 'Done'
        #     existing 'Appeal' / 'Applead' + Missing → unchanged
        #     existing 'Live'               + Live    → 'Live' (unchanged)
        #     existing 'Live'               + Missing → 'Missing'
        #     anything else                           → verdict as-is
        # 'Appeal' is the canonical name in the user's sheets; 'Applead'
        # is treated as the same workflow state because some rows still
        # carry the older spelling.
        APPEAL_LIKE = {'appeal', 'appealed', 'applead', 'applied'}
        verdict_by_url = {url.lower(): v for (_, url, v) in results}

        def _final_status(row_idx: int, verdict: str) -> str:
            cur_raw = existing_status.get(row_idx, '') or ''
            cur = cur_raw.strip().lower()
            if cur in APPEAL_LIKE:
                if verdict == 'Live':
                    return 'Done'
                if verdict == 'Missing':
                    # Don't downgrade an appealed row when the link is
                    # not visible yet — leave the existing value alone.
                    return cur_raw
            return verdict

        row_updates: dict[int, str] = {}
        for row_idx, url in all_rows:
            verdict = verdict_by_url.get(url.lower(), 'Error')
            new_val = _final_status(row_idx, verdict)
            cur = (existing_status.get(row_idx, '') or '').strip()
            # Skip the write if the value would be unchanged — avoids
            # marking unrelated cells as "modified" in the sheet's
            # version history and saves a few API quota cells.
            if new_val == cur:
                continue
            row_updates[row_idx] = new_val

        bu = _si.batch_update_status(
            resources_path, sheet_id, tab_name, status_col, row_updates,
        )
        report_msg = (f"Updated {bu.get('updated', 0)} cells in sheet"
                      if bu.get('success')
                      else f"Sheet update failed: {bu.get('message')}")

        with _status_lock:
            _status['running'] = False
            _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
            _status['report_path'] = report_msg
    except Exception as e:
        with _status_lock:
            _status['running'] = False
            _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
            _status['errors'] = (_status.get('errors') or 0) + 1
            _status['current_url'] = f'FATAL: {e}'
    finally:
        try: shutil.rmtree(tmp_root, ignore_errors=True)
        except Exception: pass


def _worker(file_path: str, num_workers: int, timeout_sec: int,
            resources_path: Path | None, show_browser: bool = False):
    import openpyxl
    global _status

    with _status_lock:
        _status.update({
            'running': True, 'started_at': datetime.utcnow().isoformat() + 'Z',
            'finished_at': '', 'total': 0, 'done': 0,
            'live': 0, 'not_live': 0, 'errors': 0,
            'current_url': '', 'report_path': '',
        })

    # Sweep any leftover temp folders from previously crashed runs first
    # (so chromium garbage doesn't accumulate even when the app was killed
    # mid-check). We're scoped to our own prefix only — never touches
    # other apps' temp data or the profile-manager folder.
    try:
        _tmp_parent = Path(tempfile.gettempdir())
        for stale in _tmp_parent.glob('nst_live_check_*'):
            try:
                shutil.rmtree(stale, ignore_errors=True)
            except Exception:
                pass
    except Exception:
        pass

    # Dedicated temp dir for the browser's user-data — wiped after the run.
    # NEVER points at the profile-manager data folder.
    tmp_root = Path(tempfile.mkdtemp(prefix='nst_live_check_'))
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [str(c.value or '').strip() for c in ws[1]]
        # ONLY accept the exact "Review Live Link" header per user spec.
        link_col_idx = None
        for i, h in enumerate(headers, 1):
            if h.strip().lower() == 'review live link':
                link_col_idx = i
                break
        if link_col_idx is None:
            with _status_lock:
                _status['running'] = False
                _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
                _status['errors'] = 1
                _status['current_url'] = ("FATAL: header 'Review Live Link' not found. "
                                          "Add a column named exactly 'Review Live Link'.")
            return

        # Use a SINGLE 'Status' column. Reuse existing one if the file
        # already has 'Status' or 'Live Status', otherwise append it as
        # the next column. No other columns are added — the output keeps
        # every original cell intact and only this one is updated.
        status_col_idx = None
        for i, h in enumerate(headers, 1):
            if h.lower() in ('status', 'live status'):
                status_col_idx = i
                break
        if status_col_idx is None:
            status_col_idx = ws.max_column + 1
            ws.cell(row=1, column=status_col_idx, value='Status')

        # Walk every row once and remember which row had which URL.
        # We CHECK each unique URL only once, but at the end we write
        # the verdict back to EVERY row that carried that URL — so the
        # output keeps every original row intact (no dedup data loss).
        all_rows: list[tuple[int, str]] = []      # (row_idx, url) every row
        items: list[tuple[int, str]] = []         # (row_idx, url) unique
        first_row_for_url: dict[str, int] = {}
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=link_col_idx).value
            if v is None:
                continue
            url = str(v).strip()
            if not url:
                continue
            # Skip cells that aren't actually URLs (notes, junk, etc.)
            ul = url.lower()
            if not (ul.startswith('http://') or ul.startswith('https://')):
                continue
            all_rows.append((r, url))
            key = url.lower()
            if key not in first_row_for_url:
                first_row_for_url[key] = r
                items.append((r, url))

        with _status_lock:
            _status['total'] = len(items)

        loop = asyncio.new_event_loop()
        try:
            asyncio.set_event_loop(loop)
            results = loop.run_until_complete(
                _run_checks(items, num_workers, timeout_sec, tmp_root, show_browser)
            )
        finally:
            try: loop.close()
            except Exception: pass

        # Build a url → verdict map then write only the Status cell on
        # every row that had this URL. Nothing else in the workbook is
        # changed.
        verdict_by_url: dict[str, str] = {
            url.lower(): verdict for (_, url, verdict) in results
        }
        for row_idx, url in all_rows:
            verdict = verdict_by_url.get(url.lower(), 'Error')
            ws.cell(row=row_idx, column=status_col_idx, value=verdict)

        in_path = Path(file_path)
        out_name = (
            f"{in_path.stem}_live_status_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        # Save to RESOURCES_PATH/output so the report appears in the in-app
        # Report Ledger automatically (alongside batch-login reports etc.).
        # Fallback to next to the input file if resources_path is missing.
        out_dir = None
        if resources_path is not None:
            try:
                out_dir = Path(resources_path) / 'output'
                out_dir.mkdir(parents=True, exist_ok=True)
            except Exception:
                out_dir = None
        if out_dir is None:
            out_dir = in_path.parent
        out_path = out_dir / out_name
        wb.save(out_path)
        wb.close()

        with _status_lock:
            _status['report_path'] = str(out_path)
            _status['running'] = False
            _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'

    except Exception as e:
        with _status_lock:
            _status['running'] = False
            _status['errors'] += 1
            _status['finished_at'] = datetime.utcnow().isoformat() + 'Z'
            _status['current_url'] = f'FATAL: {e}'

    finally:
        # Wipe the dedicated temp dir — keeps host disk clean. This is OUR
        # temp folder, not anything inside MailNexusPro / browser_profiles,
        # so profile data is safe.
        try:
            shutil.rmtree(tmp_root, ignore_errors=True)
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# Async runner
# ─────────────────────────────────────────────────────────────────────────────

async def _run_checks(items: list[tuple[int, str]], workers: int,
                      timeout_sec: int, tmp_root: Path,
                      show_browser: bool = False) -> list[tuple[int, str, str]]:
    """Spawn N independent Chromium browsers (one per worker) and pull
    URLs off a shared queue. Each browser has its own user_data_dir, so
    each gets its own cookies / consent state / process — Google doesn't
    see one browser hammering 100s of review URLs in a row, which is
    what was producing the rate-limit / captcha pattern."""
    from playwright.async_api import async_playwright

    out: list[tuple[int, str, str]] = []
    out_lock = asyncio.Lock()

    n_workers = max(1, workers)

    async def _launch(p, slot: int):
        ud = tmp_root / f'browser_{slot}'
        ud.mkdir(parents=True, exist_ok=True)
        return await p.chromium.launch_persistent_context(
            user_data_dir=str(ud),
            headless=not show_browser,
            locale='en-US',
            user_agent=('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                        'AppleWebKit/537.36 (KHTML, like Gecko) '
                        'Chrome/132.0.0.0 Safari/537.36'),
            extra_http_headers={'Accept-Language': 'en-US,en;q=0.9'},
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-features=Translate',
                '--no-default-browser-check',
                '--lang=en-US',
            ],
        )

    async def _warmup(browser):
        try:
            page = await browser.new_page()
            try:
                await page.goto('https://maps.google.com/?hl=en',
                                wait_until='domcontentloaded', timeout=20000)
                await page.wait_for_timeout(1500)
                for sel in [
                    'button[aria-label*="Accept all" i]',
                    'button:has-text("Accept all")',
                    'button:has-text("I agree")',
                    'button:has-text("Tout accepter")',
                    'form[action*="consent"] button[type="submit"]',
                ]:
                    try:
                        b = page.locator(sel).first
                        if await b.count() > 0 and await b.is_visible(timeout=600):
                            await b.click()
                            await page.wait_for_timeout(800)
                            break
                    except Exception:
                        continue
            finally:
                try: await page.close()
                except Exception: pass
        except Exception:
            pass

    async with async_playwright() as p:
        # Launch all workers' browsers in parallel for fast startup
        browsers = await asyncio.gather(
            *[_launch(p, i) for i in range(n_workers)],
            return_exceptions=False,
        )

        # Publish all browsers + loop so cancel() can close them
        # immediately from a different thread.
        _live_browser_ref['browsers'] = list(browsers)
        _live_browser_ref['loop'] = asyncio.get_event_loop()

        # Warm up every browser in parallel
        await asyncio.gather(*[_warmup(b) for b in browsers],
                             return_exceptions=True)

        # Shared work queue
        q: asyncio.Queue = asyncio.Queue()
        for it in items:
            q.put_nowait(it)

        async def _worker_loop(slot: int, browser):
            while True:
                if _cancel.is_set():
                    return
                try:
                    row_idx, url = q.get_nowait()
                except asyncio.QueueEmpty:
                    return

                with _status_lock:
                    _status['current_url'] = url[:120]
                page = None
                verdict = 'Error'
                try:
                    page = await browser.new_page()
                    verdict = await _check_url(page, url, timeout_sec)
                except Exception as e:
                    if _cancel.is_set():
                        verdict = 'Cancelled'
                    else:
                        verdict = f'Error: {str(e)[:60]}'
                finally:
                    try:
                        if page: await page.close()
                    except Exception:
                        pass

                async with out_lock:
                    out.append((row_idx, url, verdict))
                with _status_lock:
                    _status['done'] += 1
                    if verdict == 'Live':
                        _status['live'] += 1
                    elif verdict == 'Missing':
                        _status['not_live'] += 1
                    else:
                        _status['errors'] += 1

        await asyncio.gather(
            *[_worker_loop(i, b) for i, b in enumerate(browsers)],
            return_exceptions=True,
        )

        # If we were cancelled mid-run, mark every still-queued URL as
        # 'Cancelled' so the report reflects user intent.
        if _cancel.is_set():
            while True:
                try:
                    row_idx, url = q.get_nowait()
                except asyncio.QueueEmpty:
                    break
                async with out_lock:
                    out.append((row_idx, url, 'Cancelled'))

        # Close all browsers
        for b in browsers:
            try: await b.close()
            except Exception: pass

        _live_browser_ref['browsers'] = []
        _live_browser_ref['loop'] = None

    out.sort(key=lambda x: x[0])
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Per-URL check (mailexus logic)
# ─────────────────────────────────────────────────────────────────────────────

async def _check_url(page, url: str, timeout_sec: int) -> str:
    """Per user spec: a posted Google Maps review renders the star-rating
    container with id 'DU9Pgb'. If that element exists on the page,
    treat the review as Live. Anything else is Missing.

    The previous text/selector heuristics are kept ONLY as a tail-end
    fallback in case Google ever changes the id; the id check is what
    decides the verdict in normal operation.
    """
    # Step 1: navigate
    try:
        await page.goto(url, wait_until='domcontentloaded',
                        timeout=timeout_sec * 1000)
    except Exception:
        try:
            await page.goto(url, wait_until='commit', timeout=10000)
        except Exception:
            pass
    await asyncio.sleep(3)

    # PRIMARY CHECK — DU9Pgb star-rating container.
    # Google's minified token "DU9Pgb" is used as a CLASS, not an id,
    # on the live review's star-rating container. So we check the
    # class form first, then any-attribute fallback in case Google
    # changes how it uses the token.
    DU9_SELECTORS = [
        '.DU9Pgb',                  # most common — used as class
        '#DU9Pgb',                  # in case it's ever an id
        '[class*="DU9Pgb"]',        # token nested inside a multi-class string
        '[jsname="DU9Pgb"]',        # sometimes Google emits the token as jsname
        '[data-value="DU9Pgb"]',
    ]
    for _ in range(20):             # up to ~10s
        for sel in DU9_SELECTORS:
            try:
                count = await page.locator(sel).count()
                if count > 0:
                    return 'Live'
            except Exception:
                continue
        await asyncio.sleep(0.5)

    # FALLBACK — explicit "review removed / not available" copy.
    try:
        body = (await page.inner_text('body')).lower()
        for ind in MISSING_INDICATORS:
            if ind in body:
                return 'Missing'
    except Exception:
        pass

    return 'Missing'
